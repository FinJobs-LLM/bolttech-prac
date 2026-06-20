"""Fill missing values with "Unknown" for selected categorical columns in v2.

Edits cells in place via openpyxl so that every other column/value/format in the
workbook is left byte-untouched. A value is treated as missing when the cell is
empty (None) or contains only whitespace.
"""
import openpyxl

PATH = "data/preprocessed_data/claim_approval_feature_dataset_v2.xlsx"
TARGET_COLS = [
    "retailerName",
    "turnOnOff",
    "touchScreen",
    "smashed",
    "frontCamera",
    "backCamera",
    "frontOrBackCamera",
    "audio",
    "mic",
    "buttons",
]

wb = openpyxl.load_workbook(PATH)
ws = wb.active

# Header row -> column index map (1-based).
header = {cell.value: cell.column for cell in ws[1]}
missing = [c for c in TARGET_COLS if c not in header]
if missing:
    raise SystemExit(f"Columns not found in header: {missing}")

filled = {c: 0 for c in TARGET_COLS}
for col in TARGET_COLS:
    cidx = header[col]
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=cidx)
        v = cell.value
        if v is None or (isinstance(v, str) and v.strip() == ""):
            cell.value = "Unknown"
            filled[col] += 1

wb.save(PATH)
print("Rows (excl header):", ws.max_row - 1)
for c in TARGET_COLS:
    print(f"  {c}: filled {filled[c]}")
